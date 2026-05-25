import json
from pathlib import Path
import requests
import re

CONFIG_PATH = Path(__file__).parent / "config.json"
META_KEY = re.compile(r"^(AttributeItem(?:Userfield|SelectionType)\d+)_META_$")
ROOT_FIELDS = ["AssetRegisterName", "AssetNumber", "Description", "ShortDescription", "Status","OperatingStatus"]
INVALID_SHEET_CHARS = set(":\\/?*[]")

class T1Client:
    def __init__(self, service: str, config_path: Path = CONFIG_PATH):
        if not service:
            raise ValueError("Service name must be provided.")
        self.config_path = config_path
        self._token: str | None = None
        self.config = self._load_config()
        self.service = service
        self.svc_config = self.config.get("t1ws", {}).get(self.service, self.config.get(self.service, {}))

    def _load_config(self) -> dict:
        with self.config_path.open("r", encoding="utf-8") as f:
            return json.load(f)

    def get_token(self, force_refresh: bool = False) -> str:
        if self._token is None or force_refresh:
            token_cfg = self.svc_config.get("ep_get_token", {})
            url = self.svc_config["base_url"].rstrip("/") + "/" + token_cfg["url"].lstrip("/")
            method = token_cfg.get("method", "POST").upper()

            payload = {
                "client_id": token_cfg["client_id"],
                "client_secret": token_cfg["client_secret"],
                "grant_type": token_cfg["grant_type"],
            }
            headers = {"Content-Type": "application/x-www-form-urlencoded"}

            response = requests.request(method, url, data=payload, headers=headers, timeout=30, verify=False)
            response.raise_for_status()

            data = response.json()
            self._token = data["access_token"]
            
        return self._token

    def fetch_asset(self, asset_number: str, endpoint: str = "ep_asset_get") -> dict:
        ep = self.svc_config[endpoint]
        base = self.svc_config["base_url"].rstrip("/") + "/"
        path = ep["url"].lstrip("/")

        # Insert the service's asset register (e.g. "TP_AR") between path and id.
        asset_register = (self.svc_config.get("asset register") or self.svc_config.get("asset_register") or "").strip("/")
        register_segment = (asset_register + "/") if asset_register else ""

        url = base + path + register_segment + asset_number

        def _request(token: str) -> requests.Response:
            headers = {"Authorization": f"Bearer {token}"}
            return requests.get(url, headers=headers, timeout=30, verify=False)

        response = _request(self.get_token())

        if response.status_code == 401:
            response = _request(self.get_token(force_refresh=True))

        response.raise_for_status()
        return response.json()

    def save_asset(self, payload: dict, endpoint: str = "ep_asset_save") -> dict:
        ep = self.svc_config[endpoint]
        base = self.svc_config["base_url"].rstrip("/") + "/"
        path = ep["url"].lstrip("/")
        url = base + path
        method = ep.get("method", "POST").upper()

        def _request(token: str) -> requests.Response:
            headers = {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            }
            return requests.request(method, url, json=payload, headers=headers, timeout=30, verify=False)

        response = _request(self.get_token())

        if response.status_code == 401:
            response = _request(self.get_token(force_refresh=True))

        response.raise_for_status()
        return response.json() if response.content else {}

    @staticmethod
    def _infer_data_type(value) -> str:
        # 'N' for numeric (int/float), 'A' for everything else (alpha/string/null/bool).
        if isinstance(value, bool):
            return "A"
        if isinstance(value, (int, float)):
            return "N"
        return "A"

    @staticmethod
    def parse_assetitem_meta(asset: dict) -> dict:
        """Hierarchical schema:
            {
              "fields": { <fieldName>: <dataType>, ... },
              "attributes": {
                <AttributeCode>: {
                  "dataType": <char>,                       # primary SearchPath dataType
                  "levels": {                               # optional, only if captioned
                    "<n>": {
                      "<suffix>": [<caption>, <dataType>],
                      ...
                    },
                    ...
                  }
                },
                ...
              }
            }
        """
        fields: dict = {}
        attributes: dict = {}

        # Direct/root fields.
        for field in ROOT_FIELDS:
            if field in asset:
                fields[field] = T1Client._infer_data_type(asset[field])

        # Group AssetAttributes by AttributeCode (preserve first-seen order).
        groups: dict[str, list] = {}
        for entry in asset.get("AssetAttributes", []):
            code = entry.get("AttributeCode")
            if not code:
                continue
            groups.setdefault(code, []).append(entry)

        for code, entries in groups.items():
            attr_node: dict = {}

            # Primary entry → top-level dataType for this AttributeCode.
            primary = next((e for e in entries if e.get("IsPrimaryValue")), None)
            if primary is not None:
                attr_node["dataType"] = T1Client._infer_data_type(primary.get("SearchPath", ""))
            else:
                attr_node["dataType"] = "A"

            # Captioned sub-fields, grouped by integer level.
            levels: dict = {}
            for entry in entries:
                search_path = entry.get("SearchPath", "")
                level = len(search_path.split("\\")) if search_path else 0
                for key, meta_str in entry.items():
                    m = META_KEY.match(key)
                    if not m or not isinstance(meta_str, str):
                        continue
                    value_key = m.group(1)
                    if value_key not in entry:
                        continue
                    try:
                        meta = json.loads(meta_str)
                    except json.JSONDecodeError:
                        continue
                    caption = meta.get("Caption")
                    if not caption:
                        continue
                    suffix = value_key.removeprefix("AttributeItem")
                    level_dict = levels.setdefault(str(level), {})
                    if suffix not in level_dict:
                        level_dict[suffix] = [caption, meta.get("DataType", "")]

            if levels:
                # Sort levels by integer for stable column order downstream.
                attr_node["levels"] = {k: levels[k] for k in sorted(levels.keys(), key=int)}

            attributes[code] = attr_node

        return {"fields": fields, "attributes": attributes}

    def parse_assets_meta(self) -> dict:
        """
        Goes through all items under 'asset_classes' in the config,
        fetches each asset, parses its metadata, and saves the result
        to <service>.json next to config.json.
        """
        print("Fetching metadata from API for all configured asset types...")
        lookup = {}
        items = self.svc_config.get("asset_classes", [])

        for item in items:
            name = item.get("class")
            test_id = item.get("seed")
            if not name or not test_id:
                continue
            print(f"  -> Processing node: '{name}'")
            asset = self.fetch_asset(test_id)
            lookup[name] = self.parse_assetitem_meta(asset)

        meta_path = self.config_path.parent / f"{self.service}.json"
        with meta_path.open("w", encoding="utf-8") as f:
            json.dump(lookup, f, indent=2, ensure_ascii=False)
        print(f"Saved parsed metadata to {meta_path}")

        return lookup

    @staticmethod
    def _sanitize_sheet_name(name: str) -> str:
        cleaned = "".join("_" if ch in INVALID_SHEET_CHARS else ch for ch in name)
        return cleaned[:31] or "Sheet"

    @staticmethod
    def _unique_sheet_name(wb, base_name: str) -> str:
        """Sanitized sheet name; if it already exists, append '01', '02', ..."""
        name = T1Client._sanitize_sheet_name(base_name)
        if name not in wb.sheetnames:
            return name
        stem = name[:29]  # leave room for 2-digit suffix
        for i in range(1, 100):
            candidate = f"{stem}{i:02d}"
            if candidate not in wb.sheetnames:
                return candidate
        raise ValueError(f"Could not allocate unique sheet name for {base_name!r}")

    @staticmethod
    def _build_meta_columns(node_meta: dict) -> list[tuple[str, str, str, str, str, str]]:
        """Produce 6-row column tuples (kind, AttributeCode, level, suffix, dataType, header)
        from the hierarchical node_meta dict (see parse_assetitem_meta).
        Column ordering: all direct fields first, then each attribute (top-level scalar,
        followed by its captioned sub-fields sorted by level)."""
        columns: list[tuple[str, str, str, str, str, str]] = []

        # 1. Direct fields.
        for field_name, data_type in node_meta.get("fields", {}).items():
            columns.append(("", "", "", "", str(data_type), str(field_name)))

        # 2. Attributes.
        for attr_code, attr_node in node_meta.get("attributes", {}).items():
            if not isinstance(attr_node, dict):
                continue
            data_type = str(attr_node.get("dataType", "A"))
            columns.append(("Attribute", "", "", "", data_type, str(attr_code)))

            levels = attr_node.get("levels", {})
            for level_key in sorted(levels.keys(), key=lambda k: int(k) if str(k).isdigit() else 0):
                level_dict = levels[level_key]
                if not isinstance(level_dict, dict):
                    continue
                for suffix, leaf in level_dict.items():
                    caption, leaf_dt = "", ""
                    if isinstance(leaf, list) and len(leaf) >= 2:
                        caption, leaf_dt = str(leaf[0]), str(leaf[1])
                    elif isinstance(leaf, list) and len(leaf) == 1:
                        caption = str(leaf[0])
                    columns.append(("Attribute", str(attr_code), f"level_{level_key}",
                                    str(suffix), leaf_dt, caption))

        return columns

    def save_meta_to_excel(self, file: str | Path, meta: dict | None = None) -> Path:
        """Build a spreadsheet from the meta lookup.
        If `meta` is None, loads from <service>.json next to this file."""
        from openpyxl import Workbook, load_workbook  # lazy import
        from openpyxl.utils import get_column_letter  # lazy import

        if meta is None:
            meta_path = Path(__file__).parent / f"{self.service}.json"
            with meta_path.open("r", encoding="utf-8") as f:
                meta = json.load(f)

        xlsx_path = Path(file)
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)

        if xlsx_path.exists():
            wb = load_workbook(xlsx_path)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        for node_name, node_meta in meta.items():
            sheet_name = self._unique_sheet_name(wb, node_name)
            ws = wb.create_sheet(title=sheet_name)
            for col_idx, col_data in enumerate(self._build_meta_columns(node_meta), start=1):
                ws.cell(row=1, column=col_idx, value=col_data[0])
                ws.cell(row=2, column=col_idx, value=col_data[1])
                ws.cell(row=3, column=col_idx, value=col_data[2])
                ws.cell(row=4, column=col_idx, value=col_data[3])
                ws.cell(row=5, column=col_idx, value=col_data[4])
                ws.cell(row=6, column=col_idx, value=col_data[5])

                fmt = {"N": "General", "D": "yyyy-mm-dd"}.get(col_data[4], "@")
                ws.column_dimensions[get_column_letter(col_idx)].number_format = fmt

        if not wb.sheetnames:
            wb.create_sheet(title="Sheet")

        wb.save(xlsx_path)
        print(f"Saved spreadsheet to {xlsx_path}")
        return xlsx_path

    @staticmethod
    def _extract_value(asset: dict, attr_code: str, level: str, suffix: str, header: str):
        """Pull a single cell value out of a fetched asset using the column's
        5-row metadata (AttributeCode, level, suffix, dataType, header)."""
        # Captioned attribute — attr_code + suffix populated.
        if attr_code and suffix:
            target_level = 0
            if level.startswith("level_"):
                try:
                    target_level = int(level[len("level_"):])
                except ValueError:
                    pass
            value_key = "AttributeItem" + suffix
            for entry in asset.get("AssetAttributes", []):
                if entry.get("AttributeCode") != attr_code:
                    continue
                sp = entry.get("SearchPath", "")
                entry_level = len(sp.split("\\")) if sp else 0
                if entry_level != target_level:
                    continue
                if value_key in entry:
                    return entry[value_key]
            return None

        # Root field (top-level on the asset payload).
        if header in ROOT_FIELDS:
            return asset.get(header)

        # AttributeCode top-level scalar (LOCATION, SERVICEAREA, ASSET_TYPE, ...).
        for entry in asset.get("AssetAttributes", []):
            if entry.get("AttributeCode") == header and entry.get("IsPrimaryValue"):
                return entry.get("SearchPath", "")
        return None

    @staticmethod
    def is_online(file) -> bool:
        """True if `file` looks like an http(s) URL.
        Python's T1Client uses openpyxl which only handles local files;
        callers can use this to validate and fall back to a download step."""
        return isinstance(file, str) and file.lower().startswith("http")

    @staticmethod
    def _odbc_conn_str(net_conn_str: str) -> str:
        """Adapt a .NET-style connection string for pyodbc.
        Adds a DRIVER prefix if missing; the rest of the keywords are
        accepted by the SQL Server ODBC driver as-is."""
        upper = net_conn_str.upper()
        if upper.startswith("DRIVER=") or ";DRIVER=" in upper:
            return net_conn_str
        return f"DRIVER={{ODBC Driver 17 for SQL Server}};{net_conn_str}"

    @staticmethod
    def _extract_geometry_to_db(asset: dict, asset_number: str, conn, table: str) -> None:
        """Extract AssetMap.MapLayers[0] POINT geometry from `asset` and upsert
        into `table` keyed by `compkey = asset_number`, with WKT in column `wkt`."""
        asset_map = asset.get("AssetMap")
        if not isinstance(asset_map, dict):
            return
        layers = asset_map.get("MapLayers")
        if not isinstance(layers, list) or not layers:
            return
        first = layers[0]
        if str(first.get("GeometryType", "")).upper() != "POINT":
            return
        points = first.get("Points")
        if not isinstance(points, list):
            return

        coords: list[tuple[float, float]] = []
        for pt in points:
            loc = pt.get("PointLocation") if isinstance(pt, dict) else None
            if not isinstance(loc, dict):
                continue
            try:
                lat = float(loc["Latitude"])
                lon = float(loc["Longitude"])
            except (KeyError, TypeError, ValueError):
                continue
            coords.append((lat, lon))
        if not coords:
            return

        # WKT uses (longitude latitude) order.
        if len(coords) == 1:
            wkt = f"POINT ({coords[0][1]} {coords[0][0]})"
        else:
            wkt = "MULTIPOINT (" + ", ".join(f"({lon} {lat})" for lat, lon in coords) + ")"

        cur = conn.cursor()
        cur.execute(f"DELETE FROM {table} WHERE compkey = ?", asset_number)
        cur.execute(
            f"INSERT INTO {table} (compkey, wkt, sp_geometry) "
            "VALUES (?, ?, geometry::STGeomFromText(?, 4326))",
            asset_number, wkt, wkt)
        conn.commit()

    @staticmethod
    def _set_attribute_value(asset: dict, attr_code: str, level: str, suffix: str, value) -> None:
        """Mutate asset['AssetAttributes'] in place: overwrite AttributeItem<suffix>
        on the entry matching attr_code + level."""
        if not level.startswith("level_"):
            return
        try:
            target_level = int(level[len("level_"):])
        except ValueError:
            return
        value_key = "AttributeItem" + suffix
        for entry in asset.get("AssetAttributes", []):
            if entry.get("AttributeCode") != attr_code:
                continue
            sp = entry.get("SearchPath", "")
            entry_level = len(sp.split("\\")) if sp else 0
            if entry_level != target_level:
                continue
            if value_key in entry:
                entry[value_key] = value
                return

    def sync_asset_from_excel(self, file: str | Path, sheet: str, first_row: int, last_row: int, dryrun: bool = False) -> Path:
        """Sync a sheet to T1: update existing assets, create new ones for blank rows.

        For each row in [first_row, last_row] of the named sheet:
        - If column 2 (AssetNumber) is a non-empty string → fetch that asset.
        - If column 2 is blank → use the *seed* asset (svc_config['asset_classes'])
          as the create payload (patched with AssetRegisterName /
          TemplateAssetNumberInternal / AssetNumber=None), POST to ep_asset_create,
          then build the update payload from the original seed JSON with seed_id
          string-replaced by the new AssetNumber so cross-refs get retargeted.
        - Apply each non-blank cell value:
            * row-1 = 'Attribute' with level_N → mutate AssetAttributes.
            * row-1 blank → overwrite the top-level field named in row 6
              (AssetRegisterName is forced from svc_config below, not cells).
            * ASSET_TYPE columns are highlighted yellow if the cell doesn't match
              the sheet-derived class name, and forced to that class name.
        - Force AssetRegisterName from svc_config.
        - POST the modified JSON via save_asset().

        When `dryrun=True`: skip ep_asset_create and ep_asset_save POSTs and write
        the constructed payload to c:\\temp\\payload.txt for review."""
        import os
        from openpyxl import load_workbook  # lazy import
        from openpyxl.styles import PatternFill

        xlsx_path = Path(file)
        wb = load_workbook(xlsx_path)

        sheet_name = self._sanitize_sheet_name(sheet)
        if sheet_name not in wb.sheetnames:
            print(f"  -> Sheet {sheet_name} not found in workbook.")
            return xlsx_path

        # Sheet name like "Tree_Street Tree" → class name "Tree/Street Tree"
        # (first underscore becomes '/'); used to look up asset_classes config.
        underscore_idx = sheet.find("_")
        true_asset_type = (sheet[:underscore_idx] + "/" + sheet[underscore_idx + 1:]) if underscore_idx >= 0 else sheet

        ws = wb[sheet_name]
        max_col = ws.max_column

        headers = [
            (
                str(ws.cell(row=1, column=col).value or ""),  # kind
                str(ws.cell(row=2, column=col).value or ""),  # AttributeCode
                str(ws.cell(row=3, column=col).value or ""),  # level
                str(ws.cell(row=4, column=col).value or ""),  # suffix
                str(ws.cell(row=6, column=col).value or ""),  # header
            )
            for col in range(1, max_col + 1)
        ]

        asset_num_col = None
        asset_reg_col = None
        for idx, h in enumerate(headers, start=1):
            if h[4] == "AssetNumber":
                asset_num_col = idx
            elif h[4] == "AssetRegisterName":
                asset_reg_col = idx

        asset_register = self.svc_config.get("asset_register") or self.svc_config.get("asset register")

        template_id = None
        seed_id = None
        asset_classes = self.svc_config.get("asset_classes", {})
        if isinstance(asset_classes, dict):
            for class_name, class_cfg in asset_classes.items():
                if class_name.lower() == true_asset_type.lower():
                    template_id = class_cfg.get("template")
                    seed_id = class_cfg.get("seed")
                    break

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for row in range(first_row, last_row + 1):
            try:
                asset_number_val = ws.cell(row=row, column=2).value
                asset_number = asset_number_val.strip() if isinstance(asset_number_val, str) else ""

                if asset_number:
                    print(f"  -> {sheet_name} row {row}: updating asset {asset_number}")
                    asset = self.fetch_asset(asset_number)
                else:
                    if not template_id:
                        ws.cell(row=row, column=27, value=f"Missing 'template' for class '{true_asset_type}'.")
                        continue
                    if not seed_id:
                        ws.cell(row=row, column=27, value=f"Missing 'seed' for class '{true_asset_type}'.")
                        continue

                    print(f"  -> {sheet_name} row {row}: creating asset from template {template_id}")

                    # Fetch the seed; snapshot raw JSON for later seed_id replacement.
                    seed_asset = self.fetch_asset(seed_id)
                    seed_raw_str = json.dumps(seed_asset)

                    # Patch seed in place for the create call.
                    seed_asset["AssetRegisterName"] = asset_register
                    seed_asset["TemplateAssetNumberInternal"] = template_id
                    seed_asset["AssetNumber"] = None

                    if not dryrun:
                        result = self.save_asset(seed_asset, endpoint="ep_asset_create")
                        new_asset_number = result.get("AssetNumber") if isinstance(result, dict) else None
                        if not new_asset_number:
                            ws.cell(row=row, column=27, value="Create returned no AssetNumber.")
                            continue
                        new_asset_register = (result.get("AssetRegisterName") if isinstance(result, dict) else None) or asset_register
                        if asset_num_col:
                            ws.cell(row=row, column=asset_num_col, value=new_asset_number)
                        if asset_reg_col and new_asset_register:
                            ws.cell(row=row, column=asset_reg_col, value=new_asset_register)
                    else:
                        new_asset_number = f"DRYRUN_NEW_ROW_{row}"
                        new_asset_register = asset_register

                    # Build the update payload from the original seed with seed_id retargeted.
                    asset = json.loads(seed_raw_str.replace(seed_id, new_asset_number))
                    asset["AssetNumber"] = new_asset_number
                    if new_asset_register:
                        asset["AssetRegisterName"] = new_asset_register

                for col_idx, (kind, code, level, suffix, header) in enumerate(headers, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell_value = cell.value

                    # Only the top-level ASSET_TYPE column (where row-6 header == "ASSET_TYPE")
                    # gets forced to true_asset_type. Captioned attributes have code == "ASSET_TYPE"
                    # but their headers are captions ("Near Power Line", "Height(m)", ...).
                    is_asset_type = header.lower() in ("asset_type", "assettype")
                    if is_asset_type:
                        cell_value_str = str(cell_value) if cell_value is not None else ""
                        if cell_value_str.lower() != true_asset_type.lower():
                            cell.fill = yellow_fill
                        cell_value = true_asset_type

                    if cell_value is None or cell_value == "":
                        continue

                    if kind == "Attribute":
                        if code and suffix and level.startswith("level_"):
                            T1Client._set_attribute_value(asset, code, level, suffix, cell_value)
                    elif header:
                        if header.lower() == "assetregistername":
                            continue
                        asset[header] = cell_value

                if asset_register:
                    asset["AssetRegisterName"] = asset_register

                if dryrun:
                    os.makedirs(r"c:\temp", exist_ok=True)
                    dump_path = r"c:\temp\payload.txt"
                    with open(dump_path, "w", encoding="utf-8") as f:
                        json.dump(asset, f, indent=2, ensure_ascii=False)
                    ws.cell(row=row, column=27, value=f"Dry run: Payload saved to {os.path.basename(dump_path)}")
                else:
                    self.save_asset(asset)
                    ws.cell(row=row, column=27, value="")
            except Exception as e:
                ws.cell(row=row, column=27, value=str(e))

        wb.save(xlsx_path)
        print(f"Synced spreadsheet at {xlsx_path}")
        return xlsx_path

    def extract_asset(self, file: str | Path, sheet: str, first_row: int, last_row: int,
                      database_instance: str | None = None) -> Path:
        """Populate spreadsheet rows of one sheet with live asset values.

        For each row in [first_row, last_row] of the named sheet, read the
        AssetNumber, fetch the asset, and write a value into every column
        based on its 6-row header tuple (row 1 = kind).

        If `database_instance` is given and `config['database'][database_instance]`
        has `connection_string` + `table`, also extract MapLayers[0] POINT
        geometry to the configured table (compkey, wkt) — one row per asset."""
        from openpyxl import load_workbook  # lazy import

        xlsx_path = Path(file)
        wb = load_workbook(xlsx_path)

        sheet_name = self._sanitize_sheet_name(sheet)
        if sheet_name not in wb.sheetnames:
            print(f"  -> Sheet {sheet_name} not found in workbook.")
            return xlsx_path

        ws = wb[sheet_name]
        max_col = ws.max_column

        headers = [
            (
                str(ws.cell(row=2, column=col).value or ""),
                str(ws.cell(row=3, column=col).value or ""),
                str(ws.cell(row=4, column=col).value or ""),
                str(ws.cell(row=5, column=col).value or ""),
                str(ws.cell(row=6, column=col).value or ""),
            )
            for col in range(1, max_col + 1)
        ]

        asset_num_col = None
        for idx, h in enumerate(headers, start=1):
            if h[4].lower() == "assetnumber":
                asset_num_col = idx
                break

        if not asset_num_col:
            print(f"  -> No 'AssetNumber' header found in sheet {sheet_name}.")
            return xlsx_path

        # Optional: open a SQL connection if a valid database_instance is configured.
        db_conn = None
        db_table = None
        if database_instance:
            db_cfg = self.config.get("database", {}).get(database_instance)
            if db_cfg and db_cfg.get("connection_string") and db_cfg.get("table"):
                import pyodbc  # lazy import
                db_table = db_cfg["table"]
                db_conn = pyodbc.connect(T1Client._odbc_conn_str(db_cfg["connection_string"]))
                print(f"  -> DB '{database_instance}' connected; geometry -> {db_table}")
            else:
                print(f"  -> DB instance '{database_instance}' not found / incomplete in config.database.")

        try:
            for row in range(first_row, last_row + 1):
                asset_number = ws.cell(row=row, column=asset_num_col).value
                if asset_number in (None, ""):
                    continue
                try:
                    print(f"  -> {sheet_name} row {row}: fetching asset {asset_number}")
                    asset = self.fetch_asset(str(asset_number))

                    for col_idx, (attr_code, level, suffix, _data_type, header) in enumerate(headers, start=1):
                        if header.lower() == "assetnumber":
                            continue
                        value = T1Client._extract_value(asset, attr_code, level, suffix, header)
                        if value is not None:
                            ws.cell(row=row, column=col_idx, value=value)

                    db_error = None
                    if db_conn is not None and db_table:
                        try:
                            T1Client._extract_geometry_to_db(asset, str(asset_number), db_conn, db_table)
                        except Exception as db_ex:
                            db_error = f"DB: {db_ex}"

                    ws.cell(row=row, column=27, value=db_error or "")
                except Exception as e:
                    ws.cell(row=row, column=27, value=str(e))
        finally:
            if db_conn is not None:
                try:
                    db_conn.close()
                except Exception:
                    pass

        try:
            wb.save(xlsx_path)
            print(f"Updated spreadsheet at {xlsx_path}")
        except PermissionError:
            print(f"Error: Could not save to {xlsx_path} because it is open in another program (like Excel). Please close it and try again.")
        return xlsx_path
