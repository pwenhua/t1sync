# Trans.py — Python mirror of the C# Trans class.
#
# Converts a T1 ASSET CSV "template" export into a thin, import-ready
# shape. The source CSV groups every asset across multiple rows (one ASSET
# row + N ATTRIBUTE rows); Trans thins that out (one row per asset) for
# editing and walks it back to bulk-import shape.
#
# Source CSV format (T1 download template):
#   Row 1  — system-reserved literal "FORMAT ASSET, STANDARD 1.0, …"
#   Row 2  — column header. Column A is always "LineType"; the rest are
#            direct fields (AssetRegisterName, AssetNumber, …) and the
#            attribute slot columns (AttributeCode, SearchPath, LevelNumber,
#            AssetAttributeUserfield1, AssetAttributeSelectionType1, …).
#   Row 3+ — data rows, dispatched on LineType:
#              "ASSET"     — exactly one per asset record; carries that
#                            asset's direct-field values.
#              "ATTRIBUTE" — 0..N per asset, immediately following the ASSET
#                            row. Each fills a single (AttributeCode,
#                            LevelNumber) slot of the most recent ASSET row.
#
# Trans is a pure CSV-in / CSV-out pipeline — no Excel involvement. The
# class is stateless apart from the nominated-direct-fields list loaded
# from config.json; every method takes the source CSV path as its first
# argument. Public methods:
#   parse_meta(source_csv) / save_meta_to_json(source_csv, json, node_name)
#   / save_meta_to_csv(source_csv, output_csv)
#       Source CSV → hierarchical meta dict, then on disk as JSON or as a
#       6-row-header CSV (kind / code / level / suffix / dataType / header).
#   template2thin(source_csv, output_csv, asset_type_only=False)
#       Source CSV → thin CSV. Collapses each asset's ASSET + N ATTRIBUTE
#       rows into a single row, with one column per AttributeCode (cell
#       value = that attribute's SearchPath) plus one column per nominated
#       direct field. Output is a plain CSV: one column-header row + one
#       payload row per asset; no LineType column. asset_type_only=True
#       keeps only the ASSET_TYPE attribute column.
#   thin2import(source_csv, output_csv)
#       Thin CSV → T1 bulk-import CSV. Reverses template2thin: re-adds the
#       LineType column and emits one "ASSET" row plus one "ATTRIBUTE" row
#       per non-empty AttributeCode value, in the shape T1's bulk-import
#       accepts.
#   template2flat(source_csv, output_csv)
#       Like template2thin, but also exposes the ASSET_TYPE captioned
#       sub-fields (AssetAttributeUserfield<N>, AssetAttributeSelectionType<N>)
#       as their own columns whenever any asset has a value there. Columns
#       are named "ASSET_TYPE/<level>/<suffix>". Payload rows are sorted by
#       the ASSET_TYPE SearchPath.
#   flat2import(source_csv, output_csv)
#       Reverses template2flat. Behaves like thin2import for plain attribute
#       columns; for ASSET_TYPE captioned columns it re-folds them into the
#       ATTRIBUTE row's AssetAttributeUserfield<N> / SelectionType<N> cells,
#       emitting one ATTRIBUTE row per ASSET_TYPE level with values.
#   csv2xlsx(csv_path, sheet_name)
#       Convenience: load a CSV as a worksheet in the same-named xlsx
#       (csv_path with .xlsx extension). Existing workbook is reused. If
#       the proposed `sheet_name` is already taken, "1", "2"… is appended
#       until a free name is found.
#
# CLI via python-fire (`pip install fire`). Pass the method name as the
# first argument; Fire auto-routes to the matching instance method:
#   python Trans.py template2thin    <source.csv> <output.csv> [--asset_type_only]
#   python Trans.py thin2import      <source.csv> <output.csv>
#   python Trans.py template2flat    <source.csv> <output.csv>
#   python Trans.py flat2import      <source.csv> <output.csv>
#   python Trans.py save_meta_to_csv <source.csv> <output.csv>
#   python Trans.py save_meta_to_json <source.csv> <output.json> <node_name>
#   python Trans.py csv2xlsx         <source.csv> <sheet_name>
# `python Trans.py --help` lists every method; append `-- --help` after a
# method name for its own argument help.
#
# Nominated direct fields are loaded once from config.json's top-level
# "nominated_fields" array — see Trans.from_config.

from __future__ import annotations

import csv
import json
from pathlib import Path

CONFIG_PATH = Path(__file__).parent / "config.json"


class Trans:
    def __init__(self, *nominated_fields: str):
        self._nominated_fields: list[str] = [f for f in nominated_fields if f]

    # ---- factory: read nominated_fields from config.json top-level ----
    @classmethod
    def from_config(cls, config_path: str | Path = CONFIG_PATH) -> "Trans":
        fields: list[str] = []
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            nf = cfg.get("nominated_fields")
            if isinstance(nf, list):
                fields = [s for s in nf if isinstance(s, str) and s]
        except Exception:
            pass
        return cls(*fields)

    # ---- common CSV reading ----
    @staticmethod
    def _read_csv(csv_path: str | Path) -> list[list[str]]:
        with open(csv_path, "r", encoding="utf-8", newline="") as f:
            return [row for row in csv.reader(f)]

    @staticmethod
    def _header_idx(header_line: list[str]) -> dict[str, int]:
        idx: dict[str, int] = {}
        for i, name in enumerate(header_line):
            if name and name not in idx:
                idx[name] = i
        return idx

    # ---------- parse_meta (hierarchical) ----------

    def parse_meta(self, source_csv_path: str | Path) -> dict:
        rows = self._read_csv(source_csv_path)
        if len(rows) < 2:
            return {"fields": {}, "attributes": {}}

        header_line = rows[1]
        hi = self._header_idx(header_line)
        line_type_idx = hi.get("LineType", -1)
        attr_code_idx = hi.get("AttributeCode", -1)
        level_num_idx = hi.get("LevelNumber", -1)

        fields: dict = {}
        attributes: dict = {}

        # Direct (nominated) fields from the first ASSET row.
        asset_row = next(
            (r for r in rows[2:]
             if 0 <= line_type_idx < len(r) and r[line_type_idx].upper() == "ASSET"),
            None
        )
        if asset_row is not None:
            for f in self._nominated_fields:
                if f in hi and hi[f] < len(asset_row):
                    fields[f] = self._infer_data_type(asset_row[hi[f]])

        # Attributes from ATTRIBUTE rows.
        if attr_code_idx >= 0 and level_num_idx >= 0:
            for row in rows[2:]:
                if line_type_idx < 0 or line_type_idx >= len(row):
                    continue
                if row[line_type_idx].upper() != "ATTRIBUTE":
                    continue
                if attr_code_idx >= len(row):
                    continue
                code = row[attr_code_idx]
                if not code:
                    continue
                level_str = row[level_num_idx].strip() if level_num_idx < len(row) else ""
                attr_node = attributes.setdefault(code, {"dataType": "A"})
                self._add_captioned(attr_node, row, hi, level_str, "Userfield")
                self._add_captioned(attr_node, row, hi, level_str, "SelectionType")

        return {"fields": fields, "attributes": attributes}

    @staticmethod
    def _add_captioned(attr_node: dict, row: list[str], hi: dict[str, int],
                       level_str: str, family: str) -> None:
        levels = None
        level_dict = None
        for n in range(1, 21):
            col_name = f"AssetAttribute{family}{n}"
            idx = hi.get(col_name, -1)
            if idx < 0 or idx >= len(row):
                continue
            value = row[idx]
            if not value:
                continue
            if levels is None:
                levels = attr_node.setdefault("levels", {})
                level_dict = levels.setdefault(level_str, {})
            suffix = f"{family}{n}"
            if suffix in level_dict:
                continue
            # CSV doesn't carry T1's caption — leave it blank.
            level_dict[suffix] = ["", Trans._infer_data_type(value)]

    @staticmethod
    def _infer_data_type(value: str) -> str:
        if not value:
            return "A"
        try:
            float(value)
            return "N"
        except ValueError:
            pass
        try:
            from datetime import datetime
            datetime.fromisoformat(value.replace("Z", "+00:00"))
            return "D"
        except ValueError:
            pass
        return "A"

    # ---------- save_meta_to_json ----------

    def save_meta_to_json(self, source_csv_path: str | Path, json_path: str | Path, node_name: str) -> Path:
        wrapped = {node_name: self.parse_meta(source_csv_path)}
        path = Path(json_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(wrapped, f, indent=2, ensure_ascii=False)
        return path

    # ---------- save_meta_to_csv (6-row-header CSV, no data rows) ----------

    def save_meta_to_csv(self, source_csv_path: str | Path, output_csv_path: str | Path) -> Path:
        path = Path(output_csv_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        columns = self._build_meta_columns(self.parse_meta(source_csv_path))
        with open(path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            for row_idx in range(6):
                writer.writerow([c[row_idx] for c in columns])
        return path

    # ---------- template2thin: source CSV → thin CSV ----------
    #
    # The source template stores one asset across many CSV rows (one
    # LineType=ASSET row + 0..N LineType=ATTRIBUTE rows). This method
    # collapses that into a single row per asset:
    #
    #   <AttributeCode 1> … <AttributeCode N> | AssetRegisterName, AssetNumber, …
    #   ─── one column per distinct code ───    ─── nominated direct fields ───
    #
    # The cell under each AttributeCode column holds that attribute's
    # SearchPath. Captioned sub-fields (Userfield1, SelectionType2, …) are
    # NOT exploded into columns — this is the "brief" view.
    #
    # Output is a plain CSV: row 1 = column header (AttributeCode names
    # followed by nominated direct field names), row 2+ = one payload row
    # per asset. No LineType column — thin2import re-adds it.
    # asset_type_only=True keeps only the ASSET_TYPE attribute column
    # (case-insensitive).

    def template2thin(self, source_csv_path: str | Path, output_csv_path: str | Path,
                       asset_type_only: bool = False) -> Path:
        assets, attr_codes_ordered = self._read_brief(source_csv_path)

        if asset_type_only:
            attr_codes_ordered = [c for c in attr_codes_ordered
                                  if c.lower() == "asset_type"]

        # Column layout: AttributeCode columns lead (one per distinct code from
        # ATTRIBUTE rows, regardless of LevelNumber); then nominated direct
        # fields, with AssetRegisterName and AssetNumber forced to the front.
        brief_cols: list[tuple[str, str, str, str, str, str]] = []
        for code in attr_codes_ordered:
            brief_cols.append(("Attribute", "", "", "", "A", code))
        for f in self._order_nominated_fields(self._nominated_fields):
            brief_cols.append(("", "", "", "", "A", f))

        path = Path(output_csv_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            # Single column-header row.
            writer.writerow([c[5] for c in brief_cols])
            # One payload row per asset.
            for asset in assets:
                data_row: list[str] = []
                for kind, _code, level, _suffix, _dt, header in brief_cols:
                    val = ""
                    if kind == "" and header:
                        val = asset["fields"].get(header, "") or ""
                    elif kind == "Attribute" and not level:
                        val = asset["attributes"].get(header, "") or ""
                    data_row.append(val)
                writer.writerow(data_row)
        return path

    # ---------- template2flat: source CSV → flat CSV with ASSET_TYPE captions ----------
    #
    # Same shape as template2thin, plus extra columns for the ASSET_TYPE
    # captioned sub-fields whenever any asset has a non-empty value at some
    # (level, suffix). Those columns sit right after the ASSET_TYPE scalar
    # column and are named "ASSET_TYPE/<level>/<suffix>", e.g.
    # "ASSET_TYPE/1/Userfield1". Payload rows are sorted by the ASSET_TYPE
    # SearchPath (case-insensitive, ascending; empty last).

    def template2flat(self, source_csv_path: str | Path, output_csv_path: str | Path) -> Path:
        assets, attr_codes_ordered = self._read_brief(source_csv_path)

        # (level:int, suffix:str) pairs that have a non-empty value somewhere.
        slot_set: set[tuple[int, str]] = set()
        for a in assets:
            for level_str, suffix_dict in a["asset_type_captions"].items():
                if not level_str.isdigit():
                    continue
                level = int(level_str)
                for suffix, val in suffix_dict.items():
                    if val:
                        slot_set.add((level, suffix))
        caption_slots = sorted(slot_set, key=lambda s: (s[0], s[1].lower()))

        # Preserve the source's case for ASSET_TYPE; fall back to literal.
        asset_type_code = next(
            (c for c in attr_codes_ordered if c.lower() == "asset_type"),
            "ASSET_TYPE")
        has_asset_type = any(c.lower() == "asset_type" for c in attr_codes_ordered)

        # cols: list of (kind, header, code, level_str, suffix)
        cols: list[tuple[str, str, str, str, str]] = []
        if has_asset_type:
            cols.append(("Attr", asset_type_code, asset_type_code, "", ""))
            for lvl, sfx in caption_slots:
                cols.append(("Caption", f"{asset_type_code}/{lvl}/{sfx}", asset_type_code, str(lvl), sfx))
        for code in attr_codes_ordered:
            if code.lower() == "asset_type":
                continue
            cols.append(("Attr", code, code, "", ""))
        for fld in self._order_nominated_fields(self._nominated_fields):
            cols.append(("Field", fld, "", "", ""))

        # Sort assets by ASSET_TYPE SearchPath; empty last; ties on input order.
        indexed = [(i, a, (a["attributes"].get(asset_type_code, "") or "")) for i, a in enumerate(assets)]
        indexed.sort(key=lambda t: (t[2] == "", t[2].lower(), t[0]))
        sorted_assets = [t[1] for t in indexed]

        path = Path(output_csv_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow([c[1] for c in cols])
            for a in sorted_assets:
                row: list[str] = []
                for kind, header, code, level_str, suffix in cols:
                    if kind == "Attr":
                        row.append(a["attributes"].get(code, "") or "")
                    elif kind == "Caption":
                        lvl_dict = a["asset_type_captions"].get(level_str, {})
                        row.append(lvl_dict.get(suffix, "") or "")
                    else:  # Field
                        row.append(a["fields"].get(header, "") or "")
                writer.writerow(row)
        return path

    # ---------- thin2import: thin CSV → T1 bulk-import CSV ----------
    #
    # Reverses template2thin. Each input row holds one asset's data in a
    # thin shape (AttributeCode columns first, then direct fields); T1's
    # bulk import wants that asset split back into one ASSET row (carrying
    # the direct fields) plus one ATTRIBUTE row per non-empty AttributeCode
    # (carrying that code + its SearchPath value).
    #
    # Input header (saved-as CSV from template2thin):
    #   <AttributeCode 1> … <AttributeCode N> | AssetRegisterName, AssetNumber, …
    #   ─── leading attribute columns ───       ─── nominated direct fields ───
    #
    # Output:
    #   Row 1     — "FORMAT ASSET, STANDARD 1.0, DEFINITION $DEFAULT"
    #   Row 2     — LineType, <nominated direct fields…>, AttributeCode, SearchPath
    #   Row 3+    — per input asset:
    #                 row a:   LineType="ASSET", <direct values…>, blank, blank
    #                 row b…:  one per non-empty (non-"NULL") AttributeCode cell,
    #                          LineType="ATTRIBUTE", <blanks…>,
    #                          AttributeCode=<column name>, SearchPath=<cell value>
    #
    # Column matching is case-insensitive; the leading/nominated boundary is
    # the position of `AssetRegisterName` in the input header.

    def thin2import(self, source_csv_path: str | Path, output_csv_path: str | Path) -> Path:
        FORMAT_STR = "FORMAT ASSET, STANDARD 1.0, DEFINITION $DEFAULT"

        with open(source_csv_path, "r", encoding="utf-8-sig", newline="") as f:
            rows = [row for row in csv.reader(f)]

        has_format = bool(rows) and len(rows[0]) > 0 and rows[0][0] == FORMAT_STR
        if has_format:
            header_row = list(rows[1]) if len(rows) > 1 else []
            data_rows = rows[2:]
        else:
            header_row = list(rows[0]) if rows else []
            data_rows = rows[1:]

        # Boundary = index of AssetRegisterName in the source header
        # (case-insensitive). Everything before it is a "leading attribute
        # column"; its value becomes an extra ATTRIBUTE row per asset.
        boundary = 0
        for i, name in enumerate(header_row):
            if name and name.lower() == "assetregistername":
                boundary = i
                break

        leading_names = list(header_row[:boundary])

        # Drop leading cols; prepend LineType; append AttributeCode + SearchPath.
        if boundary > 0:
            del header_row[:boundary]
        header_row.insert(0, "LineType")
        ex_idx = len(header_row)
        ey_idx = ex_idx + 1
        header_row.append("AttributeCode")
        header_row.append("SearchPath")

        out_rows: list[list[str]] = [
            [FORMAT_STR],
            header_row,
        ]

        for src_row in data_rows:
            # Snapshot leading values BEFORE removing them from row a.
            leading_vals = [src_row[i] if i < len(src_row) else "" for i in range(boundary)]

            # Row a: drop leading cells, prepend "ASSET", pad to header width.
            row_a = list(src_row)
            if boundary > 0:
                del row_a[:min(boundary, len(row_a))]
            row_a.insert(0, "ASSET")
            while len(row_a) < len(header_row):
                row_a.append("")
            out_rows.append(row_a)

            # One row b per non-empty (non-"NULL") leading cell.
            for j, name in enumerate(leading_names):
                v = leading_vals[j]
                if not v or v.upper() == "NULL":
                    continue
                row_b = [""] * len(header_row)
                row_b[0]      = "ATTRIBUTE"
                row_b[ex_idx] = name
                row_b[ey_idx] = v
                out_rows.append(row_b)

        path = Path(output_csv_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            for row in out_rows:
                writer.writerow(row)
        return path

    # ---------- flat2import: flat CSV (with ASSET_TYPE captions) → T1 bulk-import CSV ----------
    #
    # Reverses template2flat. Same shape as thin2import for plain attribute
    # columns; captioned ASSET_TYPE columns (header pattern
    # "ASSET_TYPE/<level>/<suffix>") are re-folded into the
    # AssetAttributeUserfield<N> / SelectionType<N> cells of an ATTRIBUTE row
    # carrying LevelNumber=<level>. One ATTRIBUTE row is emitted per
    # ASSET_TYPE level that has at least one non-empty value.

    def flat2import(self, source_csv_path: str | Path, output_csv_path: str | Path) -> Path:
        FORMAT_STR = "FORMAT ASSET, STANDARD 1.0, DEFINITION $DEFAULT"

        with open(source_csv_path, "r", encoding="utf-8", newline="") as f:
            rows = [r for r in csv.reader(f)]

        has_format = bool(rows) and len(rows[0]) > 0 and rows[0][0] == FORMAT_STR
        if has_format:
            header_row = list(rows[1]) if len(rows) > 1 else []
            data_rows = rows[2:]
        else:
            header_row = list(rows[0]) if rows else []
            data_rows = rows[1:]

        # Boundary = position of AssetRegisterName; everything before it is a
        # leading attribute column (scalar code OR "<code>/<level>/<suffix>").
        boundary = 0
        for i, name in enumerate(header_row):
            if name and name.lower() == "assetregistername":
                boundary = i
                break

        leading_cols: list[tuple[str, bool, str, int, str]] = []  # (header, is_caption, code, level, suffix)
        max_userfield_n = 0
        max_selection_n = 0
        for i in range(boundary):
            h = header_row[i]
            parts = h.split("/")
            if len(parts) == 3 and parts[1].isdigit():
                code, lvl, sfx = parts[0], int(parts[1]), parts[2]
                leading_cols.append((h, True, code, lvl, sfx))
                low = sfx.lower()
                if low.startswith("userfield") and low[len("userfield"):].isdigit():
                    max_userfield_n = max(max_userfield_n, int(low[len("userfield"):]))
                elif low.startswith("selectiontype") and low[len("selectiontype"):].isdigit():
                    max_selection_n = max(max_selection_n, int(low[len("selectiontype"):]))
            else:
                leading_cols.append((h, False, h, 0, ""))

        has_captions = max_userfield_n > 0 or max_selection_n > 0

        # Output header.
        out_header: list[str] = ["LineType"]
        out_header.extend(header_row[boundary:])
        attr_code_idx = len(out_header); out_header.append("AttributeCode")
        search_path_idx = len(out_header); out_header.append("SearchPath")
        level_num_idx = -1
        userfield_idx: list[int] = [0] * (max_userfield_n + 1)
        selection_idx: list[int] = [0] * (max_selection_n + 1)
        if has_captions:
            level_num_idx = len(out_header); out_header.append("LevelNumber")
            for n in range(1, max_userfield_n + 1):
                userfield_idx[n] = len(out_header)
                out_header.append(f"AssetAttributeUserfield{n}")
            for n in range(1, max_selection_n + 1):
                selection_idx[n] = len(out_header)
                out_header.append(f"AssetAttributeSelectionType{n}")

        out_rows: list[list[str]] = [[FORMAT_STR], out_header]

        def blank_row() -> list[str]:
            return [""] * len(out_header)

        for src_row in data_rows:
            leading_vals = [src_row[i] if i < len(src_row) else "" for i in range(boundary)]

            # Row a: ASSET + direct fields.
            row_a = blank_row()
            row_a[0] = "ASSET"
            for i in range(len(header_row) - boundary):
                src_i = boundary + i
                row_a[1 + i] = src_row[src_i] if src_i < len(src_row) else ""
            out_rows.append(row_a)

            asset_type_sp = ""
            at_by_level: dict[int, dict[str, str]] = {}
            for j, col in enumerate(leading_cols):
                _h, is_caption, code, lvl, sfx = col
                v = leading_vals[j]
                if not v or v.upper() == "NULL":
                    continue
                is_asset_type = code.lower() == "asset_type"
                if is_caption and is_asset_type:
                    at_by_level.setdefault(lvl, {})[sfx] = v
                elif not is_caption and is_asset_type:
                    asset_type_sp = v
                elif not is_caption:
                    rb = blank_row()
                    rb[0] = "ATTRIBUTE"
                    rb[attr_code_idx] = code
                    rb[search_path_idx] = v
                    out_rows.append(rb)

            if at_by_level:
                for lvl in sorted(at_by_level):
                    lvl_dict = at_by_level[lvl]
                    rb = blank_row()
                    rb[0] = "ATTRIBUTE"
                    rb[attr_code_idx] = "ASSET_TYPE"
                    rb[search_path_idx] = asset_type_sp
                    if level_num_idx >= 0:
                        rb[level_num_idx] = str(lvl)
                    for sfx, val in lvl_dict.items():
                        low = sfx.lower()
                        if low.startswith("userfield") and low[len("userfield"):].isdigit():
                            n = int(low[len("userfield"):])
                            if 1 <= n <= max_userfield_n:
                                rb[userfield_idx[n]] = val
                        elif low.startswith("selectiontype") and low[len("selectiontype"):].isdigit():
                            n = int(low[len("selectiontype"):])
                            if 1 <= n <= max_selection_n:
                                rb[selection_idx[n]] = val
                    out_rows.append(rb)
            elif asset_type_sp:
                rb = blank_row()
                rb[0] = "ATTRIBUTE"
                rb[attr_code_idx] = "ASSET_TYPE"
                rb[search_path_idx] = asset_type_sp
                out_rows.append(rb)

        path = Path(output_csv_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            for row in out_rows:
                writer.writerow(row)
        return path

    # ---------- csv2xlsx: load CSV as a worksheet in the same-named xlsx ----------
    #
    # Output path is csv_path with the .xlsx extension; if the workbook
    # already exists it's reused and the CSV is appended as a new sheet.
    # If `sheet_name` is already taken, "1", "2"… is appended until a free
    # name is found.
    #
    # Each data row's first cell also gets a hyperlink pointing at the T1
    # AssetMyMaintenance page for that asset, parameterised by the row's
    # AssetRegisterName + AssetNumber values (header lookup is
    # case-insensitive). Rows missing either value are left as plain text.

    def csv2xlsx(self, csv_path: str | Path, sheet_name: str) -> Path:
        from openpyxl import Workbook, load_workbook
        from urllib.parse import quote

        url_template = (
            "https://maroondah-build.t1cloud.com/T1Default/CiAnywhere/Web/MAROONDAH-build/"
            "AssetsCore/AssetMyMaintenance?f=$ASC.ASSET.MNT&suite=CES"
            "&SK.AssetRegisterName={assetregistername}&SK.KeyedAssetNumber={assetnumber}"
        )

        src = Path(csv_path)
        xlsx_path = src.with_suffix(".xlsx")
        rows = self._read_csv(src)

        if xlsx_path.exists():
            wb = load_workbook(xlsx_path)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        actual = sheet_name
        n = 1
        while actual in wb.sheetnames:
            actual = f"{sheet_name}{n}"
            n += 1

        # Locate AssetRegisterName / AssetNumber columns in the header row.
        arn_col = -1
        an_col = -1
        if rows:
            for i, h in enumerate(rows[0]):
                if arn_col < 0 and h and h.lower() == "assetregistername":
                    arn_col = i
                if an_col < 0 and h and h.lower() == "assetnumber":
                    an_col = i

        ws = wb.create_sheet(actual)
        for r, row in enumerate(rows, start=1):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)

            if r == 1 or arn_col < 0 or an_col < 0:
                continue
            arn = row[arn_col] if arn_col < len(row) else ""
            an  = row[an_col]  if an_col  < len(row) else ""
            if not arn or not an:
                continue
            url = (url_template
                   .replace("{assetregistername}", quote(arn))
                   .replace("{assetnumber}", quote(an)))
            cell = ws.cell(row=r, column=1)
            cell.hyperlink = url
            cell.style = "Hyperlink"

        wb.save(xlsx_path)
        return xlsx_path

    # ---------- shared helpers (mirror MetaSchema / C# Trans privates) ----------

    def _read_brief(self, source_csv_path: str | Path) -> tuple[list[dict], list[str]]:
        """Walk the source CSV; group each LineType=ASSET row with its
        following LineType=ATTRIBUTE rows. Returns (assets, attr_codes_ordered).
        Each asset is {"fields": {...}, "attributes": {code: searchpath, ...},
        "asset_type_captions": {level_str: {suffix: value}}}.
        ASSET_TYPE captions are always collected; consumers that don't need
        them just ignore the field."""
        rows = self._read_csv(source_csv_path)
        assets: list[dict] = []
        attr_codes_ordered: list[str] = []
        seen: set[str] = set()
        if len(rows) < 2:
            return assets, attr_codes_ordered

        hi = self._header_idx(rows[1])
        line_type_idx = hi.get("LineType", -1)
        attr_code_idx = hi.get("AttributeCode", -1)
        search_path_idx = hi.get("SearchPath", -1)
        level_num_idx = hi.get("LevelNumber", -1)

        current: dict | None = None
        for row in rows[2:]:
            if line_type_idx < 0 or line_type_idx >= len(row):
                continue
            lt = row[line_type_idx].upper()
            if lt == "ASSET":
                current = {"fields": {}, "attributes": {}, "asset_type_captions": {}}
                for f in self._nominated_fields:
                    if f in hi and hi[f] < len(row):
                        current["fields"][f] = row[hi[f]]
                assets.append(current)
            elif lt == "ATTRIBUTE" and current is not None:
                if attr_code_idx < 0 or attr_code_idx >= len(row):
                    continue
                code = row[attr_code_idx]
                if not code:
                    continue
                sp = row[search_path_idx] if 0 <= search_path_idx < len(row) else ""
                if code not in current["attributes"]:
                    current["attributes"][code] = sp
                if code not in seen:
                    seen.add(code)
                    attr_codes_ordered.append(code)
                if code.lower() == "asset_type":
                    level_str = row[level_num_idx].strip() if 0 <= level_num_idx < len(row) else ""
                    if level_str:
                        self._collect_captions(current, row, hi, level_str, "Userfield")
                        self._collect_captions(current, row, hi, level_str, "SelectionType")

        return assets, attr_codes_ordered

    @staticmethod
    def _collect_captions(asset: dict, row: list[str], hi: dict[str, int],
                          level_str: str, family: str) -> None:
        level_dict: dict | None = None
        for n in range(1, 21):
            col = f"AssetAttribute{family}{n}"
            idx = hi.get(col, -1)
            if idx < 0 or idx >= len(row):
                continue
            value = row[idx]
            if not value:
                continue
            if level_dict is None:
                level_dict = asset["asset_type_captions"].setdefault(level_str, {})
            level_dict[f"{family}{n}"] = value

    @staticmethod
    def _order_nominated_fields(nominated: list[str]) -> list[str]:
        """Promote AssetRegisterName and AssetNumber to the front of the
        nominated direct-field list; preserve original order for the rest."""
        leading = ("assetregistername", "assetnumber")
        seen: set[str] = set()
        ordered: list[str] = []
        for want in leading:
            for f in nominated:
                if f.lower() == want and f.lower() not in seen:
                    ordered.append(f)
                    seen.add(f.lower())
                    break
        for f in nominated:
            if f.lower() not in seen:
                ordered.append(f)
                seen.add(f.lower())
        return ordered

    @staticmethod
    def _build_meta_columns(node_meta: dict) -> list[tuple]:
        """6-row column tuples (kind, code, level, suffix, dataType, header).
        Mirrors C# MetaSchema.BuildColumns."""
        columns: list[tuple] = []
        for f, dt in node_meta.get("fields", {}).items():
            columns.append(("", "", "", "", str(dt), f))
        for attr_code, attr_node in node_meta.get("attributes", {}).items():
            if not isinstance(attr_node, dict):
                continue
            dt = str(attr_node.get("dataType", "A"))
            columns.append(("Attribute", "", "", "", dt, attr_code))
            levels = attr_node.get("levels", {})
            for level_key in sorted(levels.keys(), key=lambda k: int(k) if str(k).isdigit() else 0):
                level_dict = levels[level_key]
                if not isinstance(level_dict, dict):
                    continue
                for suffix, leaf in level_dict.items():
                    caption, leaf_dt = "", ""
                    if isinstance(leaf, list) and len(leaf) >= 2:
                        caption, leaf_dt = str(leaf[0]), str(leaf[1])
                    elif isinstance(leaf, list) and len(leaf) == 1:
                        caption = str(leaf[0])
                    columns.append(("Attribute", attr_code, f"level_{level_key}", suffix, leaf_dt, caption))
        return columns


if __name__ == "__main__":
    import fire
    fire.Fire(Trans.from_config())

